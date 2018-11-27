from openpyxl import load_workbook

class Country:
    def __init__(self, name):
        self.name = name
        self.__years = dict()

    def addYear(self, year, totfert):
        self.__years[year] = totfert

    def __repr__(self):
        return "["+self.name+": "+str(self.getFertility())+"]"

    def __str__(self):
        return self.name + " with a fertility rate of " + str(self.getFertility())

    def getFertility(self, periods_back=0):
        if periods_back >= len(self.__years.keys()):
            periods_back = len(self.__years)-1
        year = sorted(self.__years.keys(), reverse=True)[periods_back]
        return self.__years[year]
        
        
wb = load_workbook(filename="fertility.xlsx")
sheet = wb['Fertility ']

countries = dict()
i = 0

# for every period listed
for row in sheet.iter_rows(min_row=3):
    country = row[0].value
    year = row[3].value
    totfert = row[5].value # total fertility
    try:
        float(totfert) # check if totfert is valid, since certain rows have ..
    except ValueError:
        continue # ignore that period and check the next
    else:
        if row[0].value not in countries: # if country hasn't been registered, register it
            countries[country] = Country(country)
        countries[country].addYear(year, totfert) # add that year's data to the country

high_fert = max(countries.values(), key= lambda c: c.getFertility() )
print("Highest fertility rate:", high_fert.name, "with a rate of:", high_fert.getFertility())

high_change = max(countries.values(), key= lambda c: ( float(c.getFertility(0)) - float(c.getFertility(1)) ) )
print("Highest change:", high_fert.name, "with an increase of", float(high_fert.getFertility(0)) - float(high_fert.getFertility(1)))

low_fert = min(countries.values(), key= lambda c: c.getFertility() )
print("Lowest fertility rate:", low_fert.name, "with a rate of:", low_fert.getFertility())
